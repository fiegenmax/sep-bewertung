#!/usr/bin/env python3
"""
Baut eine Excel-Uebersicht ueber alle Teams.
Liest die bereits generierten Bewertung_<team>.xlsx Dateien ein und stellt sie
nebeneinander. So siehst du auf einen Blick wer wo wie steht.

Usage:
    python3 build_overview.py
"""

import sys
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

sys.path.insert(0, str(Path(__file__).parent))
import evaluate_team as ev

FONT = "Calibri"
HEADER_FILL = PatternFill("solid", start_color="2F5496")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
CAT_FILL = PatternFill("solid", start_color="D9E2F3")
CAT_FONT = Font(name=FONT, bold=True, size=11)
TOTAL_FILL = PatternFill("solid", start_color="C6E0B4")
TOTAL_FONT = Font(name=FONT, bold=True, size=12)
NORMAL = Font(name=FONT, size=11)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def read_team_xlsx(xlsx_path):
    """Liest eine Bewertung-xlsx und gibt {criterion: (auto, manual, max)} zurueck."""
    if not xlsx_path.exists():
        return None
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception:
        return None
    ws = wb["Bewertung"] if "Bewertung" in wb.sheetnames else wb.active
    out = {}
    for row in range(6, 60):
        crit = ws.cell(row=row, column=ev.COL_CRITERION).value
        if not crit or "umme" in str(crit) or "GESAMT" in str(crit):
            continue
        auto = ws.cell(row=row, column=ev.COL_HEUR).value
        max_p = ws.cell(row=row, column=ev.COL_MAX).value
        manual = ws.cell(row=row, column=ev.COL_SCORE).value
        out[str(crit).strip()] = {"auto": auto, "manual": manual, "max": max_p}
    return out


def build_overview(mapping, output_path):
    # Sammle alle Daten
    team_data = []
    for entry in mapping:
        lf = entry["local_folder"]
        xlsx_path = ev.TEAMS / lf / f"Bewertung_{lf}.xlsx"
        data = read_team_xlsx(xlsx_path)
        if data is not None:
            team_data.append((lf, entry, data))
        else:
            print(f"   WARN: Keine xlsx fuer {lf}")

    if not team_data:
        print("Keine Team-xlsx-Dateien gefunden. Erst `python run_all.py` laufen lassen.")
        return

    # Alle Kriterien einsammeln, in der Reihenfolge der CATEGORIES
    criteria_order = []
    for cat_name, crit_names in ev.CATEGORIES:
        for c in crit_names:
            criteria_order.append((cat_name, c))
    for cat_name, crits in ev.MANUAL_CRITERIA:
        for name, _ in crits:
            criteria_order.append((cat_name, name))

    wb = Workbook()
    ws = wb.active
    ws.title = "Uebersicht"

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 8
    for col in range(4, 4 + len(team_data)):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Titel
    ws["A1"] = "Bewertungs-Uebersicht – alle Teams"
    ws["A1"].font = Font(name=FONT, bold=True, size=16)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(team_data))

    ws["A2"] = ("Werte: 'Deine Bewertung' aus den jeweiligen Team-xlsx Dateien. "
                "Wenn 'Deine Bewertung' noch leer/Auto ist, wird der Auto-Vorschlag angezeigt.")
    ws["A2"].font = Font(name=FONT, italic=True, color="808080")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3 + len(team_data))

    # Header
    HEADER_ROW = 4
    ws.cell(row=HEADER_ROW, column=1, value="Kategorie")
    ws.cell(row=HEADER_ROW, column=2, value="Kriterium")
    ws.cell(row=HEADER_ROW, column=3, value="Max")
    for i, (lf, entry, _) in enumerate(team_data):
        ws.cell(row=HEADER_ROW, column=4 + i, value=lf.replace("team-", ""))
    for col in range(1, 4 + len(team_data)):
        c = ws.cell(row=HEADER_ROW, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[HEADER_ROW].height = 30

    # Daten-Zeilen
    row = HEADER_ROW + 1
    last_cat = None
    cat_first_rows = {}
    cat_score_rows = {cat: [] for cat in set(c for c, _ in criteria_order)}

    for cat_name, crit in criteria_order:
        ws.cell(row=row, column=1, value=cat_name if cat_name != last_cat else "")
        if cat_name != last_cat:
            ws.cell(row=row, column=1).font = CAT_FONT
            ws.cell(row=row, column=1).fill = CAT_FILL
            cat_first_rows[cat_name] = row
            last_cat = cat_name
        ws.cell(row=row, column=2, value=crit).font = NORMAL
        # Max aus dem ersten Team mit Daten holen
        max_p = None
        for _, _, data in team_data:
            if crit in data:
                max_p = data[crit]["max"]
                break
        ws.cell(row=row, column=3, value=max_p).alignment = Alignment(horizontal="center")
        # Team-Werte
        for i, (lf, _, data) in enumerate(team_data):
            crit_data = data.get(crit, {})
            value = crit_data.get("manual")
            if value is None or (isinstance(value, str) and value.strip() == ""):
                value = crit_data.get("auto")
            # "x" Markierung: noch nicht ausgefuellt
            ws.cell(row=row, column=4 + i, value=value).alignment = Alignment(horizontal="center")
        cat_score_rows[cat_name].append(row)
        for col in range(1, 4 + len(team_data)):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    # Zwischensummen pro Kategorie (am Ende der Kategorie)
    row += 1
    ws.cell(row=row, column=1, value="GESAMT").font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = TOTAL_FILL
    # Max
    total_max = sum(rows for rows in [
        sum(data["max"] for data in tdata.values() if isinstance(data.get("max"), (int, float)))
        for _, _, tdata in team_data[:1]
    ])
    ws.cell(row=row, column=3, value=total_max).font = TOTAL_FONT
    ws.cell(row=row, column=3).fill = TOTAL_FILL
    # Pro Team: SUM aller Kriterien-Zeilen
    for i in range(len(team_data)):
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(4 + i)
        all_rows = [r for rows in cat_score_rows.values() for r in rows]
        # SUM ignoriert "x" Strings
        cell_refs = ",".join(f"{col_letter}{r}" for r in all_rows)
        ws.cell(row=row, column=4 + i, value=f"=SUM({cell_refs})").font = TOTAL_FONT
        ws.cell(row=row, column=4 + i).fill = TOTAL_FILL
    for col in range(1, 4 + len(team_data)):
        ws.cell(row=row, column=col).border = BORDER
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")

    # Conditional Formatting: Farb-Skala fuer die Team-Spalten
    from openpyxl.utils import get_column_letter
    for i in range(len(team_data)):
        col_letter = get_column_letter(4 + i)
        range_str = f"{col_letter}{HEADER_ROW+1}:{col_letter}{row-1}"
        rule = ColorScaleRule(start_type="min", start_color="F8CBAD",
                              mid_type="percentile", mid_value=50, mid_color="FFE699",
                              end_type="max", end_color="C6E0B4")
        ws.conditional_formatting.add(range_str, rule)

    ws.freeze_panes = "D5"
    wb.save(output_path)


def main():
    # Kein load_config()/Token noetig: die Uebersicht liest nur die bereits
    # generierten Bewertung_<team>.xlsx + team_mapping.json.
    mapping = json.loads((ev.OUTPUTS / "team_mapping.json").read_text(encoding="utf-8"))
    out = ev.TEAMS / "Uebersicht_alle_Teams.xlsx"
    print(f"Generiere Uebersicht ueber {len(mapping)} Teams ...")
    build_overview(mapping, out)
    print(f"OK: {out}")


if __name__ == "__main__":
    main()
