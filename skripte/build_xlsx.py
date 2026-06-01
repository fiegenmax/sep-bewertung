#!/usr/bin/env python3
"""
Excel-Bewertungsbogen aus den Daten von evaluate_team.py bauen.

Die Excel-Datei enthält:
- Sheet 'Bewertung' mit allen Kriterien, vorausgefuellten Auto-Vorschlaegen
  und einer Spalte 'Deine Bewertung' zum manuellen Eintragen
- Summen pro Kategorie + Gesamtsumme als Excel-Formeln
- Sheet 'Zusatzinfos' mit Commit-Verteilung, CI-Status etc.

Usage:
    python3 build_xlsx.py <local_team_folder>
"""

import sys
import json
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# Wiederverwendung der Analyse aus evaluate_team
sys.path.insert(0, str(Path(__file__).parent))
import evaluate_team as ev


CATEGORIES = ev.CATEGORIES
MANUAL_CRITERIA = ev.MANUAL_CRITERIA

# Spalten-Buchstaben aus den zentralen Indizes (evaluate_team.COL_*) ableiten,
# damit Schreib-Formeln und Lese-Logik nicht auseinanderdriften koennen.
L_HEUR = get_column_letter(ev.COL_HEUR)    # "C"
L_LLM = get_column_letter(ev.COL_LLM)      # "D"
L_MAX = get_column_letter(ev.COL_MAX)      # "E"
L_SCORE = get_column_letter(ev.COL_SCORE)  # "F"

# Styles
FONT = "Calibri"
HEADER_FILL = PatternFill("solid", start_color="2F5496")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
CAT_FILL = PatternFill("solid", start_color="D9E2F3")
CAT_FONT = Font(name=FONT, bold=True, size=11)
SUM_FILL = PatternFill("solid", start_color="FFF2CC")
SUM_FONT = Font(name=FONT, bold=True, size=11)
TOTAL_FILL = PatternFill("solid", start_color="C6E0B4")
TOTAL_FONT = Font(name=FONT, bold=True, size=12)
INPUT_FILL = PatternFill("solid", start_color="FFFFCC")
AUTO_FONT = Font(name=FONT, italic=True, color="808080", size=10)
NORMAL = Font(name=FONT, size=11)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def check_criterion_coverage(by_crit):
    """Gibt die CATEGORIES-Kriterien zurueck, fuer die KEIN Result vorliegt.
    Verhindert, dass eine String-/Umlaut-Drift eine Zeile still verschwinden laesst."""
    expected = [c for _, crits in CATEGORIES for c in crits]
    return [c for c in expected if c not in by_crit]


def _apply_to_row(ws, row, cols, **kwargs):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        for k, v in kwargs.items():
            setattr(cell, k, v)


def extract_manual_values(xlsx_path):
    """Liest aus einer bestehenden Bewertungs-xlsx die manuellen Eintragungen aus
    der Spalte 'Deine Bewertung' (F) und 'Anmerkungen' (G) je Kriterium.
    Schluessel ist der Kriterium-Name (Spalte B).
    Returns {criterion: {"score": x, "note": y}}.
    """
    if not xlsx_path.exists():
        return {}
    try:
        wb = load_workbook(xlsx_path, data_only=False)
        ws = wb["Bewertung"] if "Bewertung" in wb.sheetnames else wb.active
    except Exception:
        return {}
    out = {}
    # Auto-Detect: alte (7 Spalten) oder neue Struktur (9 Spalten)
    # In neuer Struktur: F=Deine Bewertung, G=Anmerkungen
    # In alter Struktur: E=Deine Bewertung, F=Anmerkungen
    header_f = ws.cell(row=5, column=ev.COL_SCORE).value
    new_layout = (header_f and "Deine" in str(header_f))
    score_col = ev.COL_SCORE if new_layout else 5
    note_col = ev.COL_NOTE if new_layout else 6
    for row in range(6, 60):
        crit = ws.cell(row=row, column=ev.COL_CRITERION).value
        if not crit or "umme" in str(crit) or "GESAMT" in str(crit):
            continue
        e = ws.cell(row=row, column=score_col).value
        f = ws.cell(row=row, column=note_col).value
        out[str(crit).strip()] = {"score": e, "note": f}
    return out


def merge_manual_values_into_workbook(wb, old_values):
    """Schreibt die alten manuellen Werte in die neu generierte Excel.
    Nur dort wo der alte Wert vom Auto-Vorschlag abweicht oder eine Notiz hat.
    """
    if not old_values:
        return 0
    ws = wb["Bewertung"] if "Bewertung" in wb.sheetnames else wb.active
    overrides = 0
    for row in range(6, 60):
        crit = ws.cell(row=row, column=ev.COL_CRITERION).value
        if not crit or "umme" in str(crit) or "GESAMT" in str(crit):
            continue
        key = str(crit).strip()
        if key not in old_values:
            continue
        old = old_values[key]
        # Score nur uebernehmen wenn er sich vom aktuellen Auto-Vorschlag (Spalte C) unterscheidet
        # oder ein "x" ist (= manueller Eintrag noch ausstehend / war schon manuell)
        auto = ws.cell(row=row, column=ev.COL_HEUR).value
        old_score = old.get("score")
        if old_score is not None and old_score != "" and old_score != auto:
            # Aber: wenn der alte Wert eine Formel ist (Zwischensumme), nicht uebernehmen
            if not (isinstance(old_score, str) and old_score.startswith("=")):
                ws.cell(row=row, column=ev.COL_SCORE).value = old_score   # F = Deine Bewertung
                overrides += 1
        old_note = old.get("note")
        if old_note and str(old_note).strip():
            ws.cell(row=row, column=ev.COL_NOTE).value = old_note  # G = Anmerkungen
            overrides += 1
    return overrides


def backup_existing(xlsx_path):
    """Macht eine .bak vor dem Ueberschreiben. Returns Pfad oder None."""
    if not xlsx_path.exists():
        return None
    bak = xlsx_path.with_suffix(".xlsx.bak")
    shutil.copy2(xlsx_path, bak)
    return bak


def build_xlsx(team_name, gitlab_path, web_url, results, output_path,
               keep_manual=True, do_backup=True):
    # Zielverzeichnis sicherstellen: die Team-Ordner unter teams/ sind gitignored
    # (Studentendaten) und fehlen auf einem frischen Checkout. Ohne dieses mkdir
    # crasht wb.save() mit FileNotFoundError, weil teams/team-<name>/ nicht existiert.
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    by_crit = {r["criterion"]: r for r in results}
    info_results = [r for r in results if r.get("max") == 0]

    missing = check_criterion_coverage(by_crit)
    if missing:
        print(f"WARN: {len(missing)} Kriterium(en) ohne Result (Drift in criterion-Strings?): "
              f"{missing}", file=sys.stderr)

    # Schritt 1: alte manuelle Werte einlesen falls vorhanden
    old_values = extract_manual_values(output_path) if keep_manual else {}
    # Schritt 2: Backup
    bak_path = backup_existing(output_path) if do_backup else None

    wb = Workbook()
    ws = wb.active
    ws.title = "Bewertung"

    # Spaltenbreite
    # A=Kategorie B=Kriterium C=Heur D=LLM E=Max F=Deine_Bew G=Anmerk H=Begr_Heur I=Begr_LLM
    widths = {"A": 32, "B": 38, "C": 9, "D": 9, "E": 6, "F": 14, "G": 32, "H": 50, "I": 50}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # === Titel ===
    ws["A1"] = f"Bewertungsbogen – {team_name}"
    ws["A1"].font = Font(name=FONT, bold=True, size=16)
    ws.merge_cells("A1:I1")

    ws["A2"] = "GitLab:"
    ws["A2"].font = Font(name=FONT, bold=True)
    ws["B2"] = gitlab_path
    ws["B2"].hyperlink = web_url
    ws["B2"].font = Font(name=FONT, color="0563C1", underline="single")
    ws.merge_cells("B2:I2")

    ws["A3"] = "Hinweis:"
    ws["A3"].font = Font(name=FONT, bold=True)
    ws["B3"] = ("Spalte C zeigt den heuristischen Auto-Vorschlag, Spalte D den LLM-Score (Zweitmeinung). "
                "'Deine Bewertung' (gelb, F) ist mit dem heuristischen Vorschlag vorausgefüllt und überschreibbar. "
                "'x' bedeutet: noch manuell auszufüllen. Rote/fette Zellen in F: Wert vom Heuristik-Vorschlag geändert. "
                "Orange in Spalte D: Heuristik und LLM weichen >1 Punkt ab – manuell prüfen. "
                "Summen aktualisieren sich automatisch. Begründungen separat: Heuristik (H) und LLM (I).")
    ws["B3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("B3:I3")
    ws.row_dimensions[3].height = 45

    # === Provenienz-Kopfzeile (Datengrundlage) ===
    prov_res = next((r for r in results if r.get("criterion") == "Datengrundlage (Info)"), None)
    if prov_res:
        p = prov_res["details"]
        ws["A4"] = "Datengrundlage:"
        ws["A4"].font = Font(name=FONT, bold=True, size=9)
        ws["B4"] = (f"Commit {p['head_sha'][:10]} ({p['branch']}) · Projekt {p['project_id']} · "
                    f"geholt {p['fetched_at']} UTC")
        ws["B4"].font = Font(name=FONT, italic=True, color="808080", size=9)
        ws.merge_cells("B4:I4")

    # === Header-Zeile ===
    HEADER_ROW = 5
    headers = ["Kategorie", "Kriterium", "Heur-Score", "LLM-Score", "Max",
               "Deine Bewertung", "Anmerkungen",
               "Begründung Heuristik", "Begründung LLM"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=HEADER_ROW, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[HEADER_ROW].height = 30

    # === Daten ===
    row = HEADER_ROW + 1
    category_sum_rows = []
    score_cells = []  # Liste der "Deine Bewertung"-Zellen für die Gesamtsumme
    criterion_rows = []  # Liste aller Kriterien-Zeilen-Nummern (fuer LLM-Hybrid-Summe)

    for cat_name, crit_names in CATEGORIES:
        first_row_of_cat = row
        cat_score_cells = []

        for crit in crit_names:
            r = by_crit.get(crit)
            if not r:
                continue
            ws.cell(row=row, column=1, value=cat_name if row == first_row_of_cat else "")
            if row == first_row_of_cat:
                ws.cell(row=row, column=1).font = CAT_FONT
                ws.cell(row=row, column=1).fill = CAT_FILL
            ws.cell(row=row, column=2, value=r["criterion"]).font = NORMAL
            # C: Heur-Score
            ws.cell(row=row, column=3, value=r["score"]).font = AUTO_FONT
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
            # D: LLM-Score (nur wenn LLM-Review vorhanden)
            llm_rev = (r.get("details") or {}).get("llm_review")
            llm_score = llm_rev["score"] if llm_rev else None
            llm_reason = llm_rev["reason"] if llm_rev else ""
            d_cell = ws.cell(row=row, column=4, value=llm_score)
            d_cell.font = Font(name=FONT, italic=True, color="6B5B95", size=10)
            d_cell.alignment = Alignment(horizontal="center")
            # E: Max
            ws.cell(row=row, column=5, value=r["max"]).font = NORMAL
            ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
            # F: Deine Bewertung
            input_cell = ws.cell(row=row, column=6, value=r["score"])
            input_cell.fill = INPUT_FILL
            input_cell.font = Font(name=FONT, bold=True, size=11)
            input_cell.alignment = Alignment(horizontal="center")
            input_cell.border = BORDER
            # G: Anmerkungen
            ws.cell(row=row, column=7, value="").fill = INPUT_FILL
            ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical="top")
            # H: Begruendung Heuristik
            ws.cell(row=row, column=8, value=r["reason"]).font = AUTO_FONT
            ws.cell(row=row, column=8).alignment = Alignment(wrap_text=True, vertical="top")
            # I: Begruendung LLM
            ws.cell(row=row, column=9, value=llm_reason).font = Font(
                name=FONT, italic=True, color="6B5B95", size=10)
            ws.cell(row=row, column=9).alignment = Alignment(wrap_text=True, vertical="top")
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = BORDER

            cat_score_cells.append(f"{L_SCORE}{row}")
            score_cells.append(f"{L_SCORE}{row}")
            criterion_rows.append(row)
            row += 1

        # Zwischensumme
        cat_max = sum(by_crit[c]["max"] for c in crit_names if c in by_crit)
        ws.cell(row=row, column=2, value=f"Zwischensumme {cat_name}").font = SUM_FONT
        if cat_score_cells:
            sum_formula = f"=SUM({','.join(cat_score_cells)})"
            ws.cell(row=row, column=3, value=sum_formula).font = SUM_FONT
            ws.cell(row=row, column=6, value=sum_formula).font = SUM_FONT
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=cat_max).font = SUM_FONT
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = SUM_FILL
            ws.cell(row=row, column=col).border = BORDER
        category_sum_rows.append(row)
        row += 1

        # Leere Trennzeile
        row += 1

    # === Manuelle Kriterien ===
    for manual_cat, crits in MANUAL_CRITERIA:
        first_row = row
        for name, max_p in crits:
            ws.cell(row=row, column=1, value=manual_cat if row == first_row else "")
            if row == first_row:
                ws.cell(row=row, column=1).font = CAT_FONT
                ws.cell(row=row, column=1).fill = CAT_FILL
            ws.cell(row=row, column=2, value=name).font = NORMAL
            ws.cell(row=row, column=3, value="").font = AUTO_FONT
            ws.cell(row=row, column=5, value=max_p).font = NORMAL
            ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
            # "x" als Default - SUM ignoriert Text, der User ersetzt es durch eine Zahl
            input_cell = ws.cell(row=row, column=6, value="x")
            input_cell.fill = INPUT_FILL
            input_cell.font = Font(name=FONT, bold=True, color="C00000", size=12)
            input_cell.alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=7, value="").fill = INPUT_FILL
            ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row, column=8, value="Nach mündlicher Prüfung manuell eintragen.").font = AUTO_FONT
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = BORDER
            score_cells.append(f"{L_SCORE}{row}")
            row += 1

        manual_max = sum(m for _, m in crits)
        ws.cell(row=row, column=2, value=f"Zwischensumme {manual_cat}").font = SUM_FONT
        manual_score_refs = [f"{L_SCORE}{r}" for r in range(first_row, row)]
        sum_formula = f"=SUM({','.join(manual_score_refs)})"
        ws.cell(row=row, column=3, value=sum_formula).font = SUM_FONT
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=manual_max).font = SUM_FONT
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=sum_formula).font = SUM_FONT
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = SUM_FILL
            ws.cell(row=row, column=col).border = BORDER
        category_sum_rows.append(row)
        row += 2

    # === Gesamtsumme ===
    total_max = sum(by_crit[c]["max"] for _, crits in CATEGORIES for c in crits if c in by_crit)
    total_max += sum(m for _, crits in MANUAL_CRITERIA for _, m in crits)
    # Spalte C (GESAMT Heuristik) summiert die heuristischen C-Scores der Kriterien,
    # NICHT die F-Werte. Spalte F summiert weiterhin "Deine Bewertung" (inkl. manuell).
    heur_cells = [f"{L_HEUR}{r}" for r in criterion_rows]
    ws.cell(row=row, column=2, value="GESAMT").font = TOTAL_FONT
    heur_total_formula = f"=SUM({','.join(heur_cells)})" if heur_cells else 0
    ws.cell(row=row, column=3, value=heur_total_formula).font = TOTAL_FONT
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=5, value=total_max).font = TOTAL_FONT
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=6, value=f"=SUM({','.join(score_cells)})").font = TOTAL_FONT
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
    for col in range(1, 10):
        ws.cell(row=row, column=col).fill = TOTAL_FILL
        ws.cell(row=row, column=col).border = BORDER
    row += 1

    # === GESAMT (LLM-Hybrid): nimmt D wo vorhanden, sonst C ===
    if criterion_rows:
        # Formel: pro Zeile IF(D{r}="",C{r},D{r}), aufsummiert
        # Wir bauen sie als Summe einzelner IF-Terme - das geht ohne Array-Formel
        terms = [f'IF({L_LLM}{r}="",{L_HEUR}{r},{L_LLM}{r})' for r in criterion_rows]
        hybrid_formula = "=" + "+".join(terms)
        ws.cell(row=row, column=2, value="GESAMT (LLM-Hybrid)").font = TOTAL_FONT
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=3, value=hybrid_formula).font = TOTAL_FONT
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        # Hinweis-Spalte
        ws.cell(row=row, column=4, value="(D wo, sonst C)").font = AUTO_FONT
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=total_max - sum(m for _, crits in MANUAL_CRITERIA for _, m in crits)).font = TOTAL_FONT
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        # Hinweis in der Begruendung
        ws.cell(row=row, column=8, value=(
            "LLM-Hybrid-Summe: pro Kriterium wird der LLM-Score (Spalte D) genommen, "
            "wo er existiert, sonst der Heuristik-Score (Spalte C). "
            "Manuelle Kriterien sind hier nicht enthalten."
        )).font = AUTO_FONT
        ws.cell(row=row, column=8).alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = PatternFill("solid", start_color="E7DFEC")
            ws.cell(row=row, column=col).border = BORDER
    row += 2

    # === Conditional Formatting: Zelle E mit rotem Rand wenn != Auto-Vorschlag (C)
    # Wir packen die Regel auf den gesamten E-Bereich der Kriterien-Zeilen
    from openpyxl.styles import Border, Side
    red_border = Border(left=Side(border_style="medium", color="C00000"),
                        right=Side(border_style="medium", color="C00000"),
                        top=Side(border_style="medium", color="C00000"),
                        bottom=Side(border_style="medium", color="C00000"))
    # Regel NUR auf die echten Kriterien-Zeilen anwenden (nicht auf Summen-/Leer-
    # zeilen, die frueher mit im Bereich F{HEADER_ROW+1}:F{row-3} lagen). criterion_rows
    # ist nicht zusammenhaengend -> in zusammenhaengende Bloecke gruppieren.
    def _contiguous_runs(rows):
        runs, start, prev = [], None, None
        for r in rows:
            if start is None:
                start = prev = r
            elif r == prev + 1:
                prev = r
            else:
                runs.append((start, prev)); start = prev = r
        if start is not None:
            runs.append((start, prev))
        return runs

    if criterion_rows:
        anchor = criterion_rows[0]  # Bezugszelle fuer die relativen Formel-Refs
        range_str = " ".join(f"{L_SCORE}{a}:{L_SCORE}{b}"
                             for a, b in _contiguous_runs(criterion_rows))
        rule = FormulaRule(
            formula=[f'AND({L_HEUR}{anchor}<>"",{L_SCORE}{anchor}<>{L_HEUR}{anchor})'],
            font=Font(name=FONT, bold=True, color="C00000", size=12),
        )
        ws.conditional_formatting.add(range_str, rule)

        # Divergenz-Flag: Spalte D faerben, wo Heuristik (C) und LLM (D) >1 Punkt auseinanderliegen.
        d_anchor = criterion_rows[0]
        d_range = " ".join(f"{L_LLM}{a}:{L_LLM}{b}" for a, b in _contiguous_runs(criterion_rows))
        diverge_rule = FormulaRule(
            formula=[f'AND({L_LLM}{d_anchor}<>"",ABS({L_LLM}{d_anchor}-{L_HEUR}{d_anchor})>1)'],
            fill=PatternFill("solid", start_color="FFD9A0"),
        )
        ws.conditional_formatting.add(d_range, diverge_rule)

    # Freeze: Header oben
    ws.freeze_panes = "A6"

    # === Sheet "Zusatzinfos" ===
    ws2 = wb.create_sheet("Zusatzinfos")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 90

    ws2["A1"] = "Zusatzinformationen"
    ws2["A1"].font = Font(name=FONT, bold=True, size=14)
    ws2.merge_cells("A1:B1")
    ws2["A2"] = "Diese Infos haben keinen direkten Score, helfen aber bei der manuellen Bewertung."
    ws2["A2"].font = Font(name=FONT, italic=True, color="808080")
    ws2.merge_cells("A2:B2")

    r2 = 4
    for info in info_results:
        ws2.cell(row=r2, column=1, value=info["criterion"]).font = Font(name=FONT, bold=True, size=11)
        ws2.cell(row=r2, column=1).fill = CAT_FILL
        ws2.cell(row=r2, column=2, value=info["reason"]).font = NORMAL
        ws2.cell(row=r2, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[r2].height = 60
        r2 += 1
        d = info.get("details", {})
        if d:
            detail_text = json.dumps(d, ensure_ascii=False, indent=2)
            ws2.cell(row=r2, column=1, value="Details").font = AUTO_FONT
            ws2.cell(row=r2, column=2, value=detail_text).font = AUTO_FONT
            ws2.cell(row=r2, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws2.row_dimensions[r2].height = max(60, min(220, detail_text.count("\n") * 15))
            r2 += 1
        r2 += 1

    # Schritt 3: alte manuelle Werte uebernehmen
    if keep_manual and old_values:
        merged = merge_manual_values_into_workbook(wb, old_values)
        if merged:
            print(f"   {merged} Wert/Eintrag aus vorheriger Bewertung uebernommen.")

    wb.save(output_path)
    if bak_path:
        print(f"   Backup: {bak_path.name}")


def main():
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <local_team_folder>")
        sys.exit(2)
    local_folder = sys.argv[1]
    cfg = ev.load_config()
    token = cfg["GITLAB_TOKEN"]
    yaml_cfg = ev.load_yaml_config()
    import llm as llm_mod
    llm_client = llm_mod.load_llm_from_configs(cfg, yaml_cfg)

    mapping = json.loads((ev.OUTPUTS / "team_mapping.json").read_text(encoding="utf-8"))
    entry = next((m for m in mapping if m["local_folder"] == local_folder), None)
    if not entry:
        print(f"FEHLER: Kein Mapping fuer {local_folder}")
        sys.exit(1)

    results = ev.collect_results(entry, token, llm_client=llm_client, yaml_cfg=yaml_cfg)

    out = ev.TEAMS / local_folder / f"Bewertung_{local_folder}.xlsx"
    build_xlsx(local_folder, entry["gitlab_path"], entry["web_url"], results, out)
    print(f"OK: Geschrieben: {out}")


if __name__ == "__main__":
    main()
