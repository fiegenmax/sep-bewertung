#!/usr/bin/env python3
"""
Befuellt das Artifacts-Exam-PDF-Formular mit Werten aus einer
generierten Bewertung_<team>.xlsx Datei.

Optional aufgerufen aus run_all.py mit --pdf Flag oder direkt:
    python3 fill_pdf.py team-entropy

Was befuellt wird (und NUR das):
  - "Geprüftes Team"  (Textfeld1)  -> Teamname
  - "Prüfer"          (Textfeld0)  -> Pruefername (PRUEFER_NAME)
  - "Gesamtpunktzahl" (Textfeld2)  -> Summe aller "Deine Bewertung"-Punkte (Spalte F)
  - Zwischensummen je Abschnitt (Textfeld3/13/17)
  - die zur jeweiligen Punktzahl gehoerenden Kontrollkaestchen

Bewusst NICHT befuellt: die Anmerkungs-/Kommentar-Textfelder (Textfeld4-22).

Feld-Namen und die Score->Checkbox-Zuordnung wurden direkt aus der Geometrie
des Templates verifiziert (siehe CHECKBOX_MAP-Kommentar). Die Zuordnung ist durch
test_evaluate.TestCheckboxMapMatchesGeometry gegen das PDF abgesichert.

"x"/nicht-numerische Werte in Spalte F (= noch nicht bewertet) werden
uebersprungen: kein Kreuz, kein Beitrag zur Summe. Spalten-Indizes kommen
aus evaluate_team.COL_*.
"""

import sys
import json
from pathlib import Path

from openpyxl import load_workbook
# pypdf wird nur fuer das tatsaechliche Befuellen (fill_pdf) gebraucht und daher
# lazy importiert - so bleibt read_xlsx_scores nutzbar, auch wenn pypdf fehlt.

sys.path.insert(0, str(Path(__file__).parent))
import evaluate_team as ev


# ---------------------------------------------------------------------------
# Mapping: Kriterium -> {score: checkbox_field_name}
#
# WICHTIG: Die VERTIKALEN Antwort-Listen im Template sind INVERS nummeriert:
# das oberste Kaestchen ("Nicht durchgeführt" = 0 Punkte) hat die HOECHSTE
# Kontrollkaestchen-ID, das unterste ("Vollständig durchgeführt") die niedrigste.
# Die HORIZONTALEN Skalen (Code-Doku 14-19, Tests 22-29, Release 30-35) sind
# dagegen aufsteigend (linkes Kaestchen = 0). Beides ist aus den Feld-Koordinaten
# verifiziert. Nicht "aufraeumen" ohne die Geometrie erneut zu pruefen.
# ---------------------------------------------------------------------------
CHECKBOX_MAP = {
    # --- Sprintdokumentation (Seite 0, vertikal -> invers) ---
    "User Stories / Issues ordentlich erstellt": {
        0: "Kontrollkästchen3", 1: "Kontrollkästchen2",
        2: "Kontrollkästchen1", 3: "Kontrollkästchen0",
    },
    "Verständliche Commit-Messages": {0: "Kontrollkästchen5", 1: "Kontrollkästchen4"},
    "Team-Meetings dokumentiert": {0: "Kontrollkästchen7", 1: "Kontrollkästchen6"},
    "Release mit Changelog/Release-Notes": {0: "Kontrollkästchen9", 1: "Kontrollkästchen8"},
    "Sinnvolle Epics + Verlinkung": {0: "Kontrollkästchen11", 1: "Kontrollkästchen10"},
    # --- Code-Qualität (Seite 1) ---
    "Code sinnvoll strukturiert": {0: "Kontrollkästchen13", 1: "Kontrollkästchen12"},
    "Code ausreichend dokumentiert": {  # horizontal -> aufsteigend
        0: "Kontrollkästchen14", 1: "Kontrollkästchen15", 2: "Kontrollkästchen16",
        3: "Kontrollkästchen17", 4: "Kontrollkästchen18", 5: "Kontrollkästchen19",
    },
    "Code sauber/ohne größere Mängel": {0: "Kontrollkästchen21", 1: "Kontrollkästchen20"},
    "Tests vorhanden und sinnvoll": {  # horizontal -> aufsteigend
        0: "Kontrollkästchen22", 1: "Kontrollkästchen23", 2: "Kontrollkästchen24",
        3: "Kontrollkästchen25", 4: "Kontrollkästchen26", 5: "Kontrollkästchen27",
        6: "Kontrollkästchen28", 7: "Kontrollkästchen29",
    },
    # --- Implementierte Funktionalität (Seite 1-2) ---
    "Release ausfuehrbar": {  # horizontal -> aufsteigend
        0: "Kontrollkästchen30", 1: "Kontrollkästchen31", 2: "Kontrollkästchen32",
        3: "Kontrollkästchen33", 4: "Kontrollkästchen34", 5: "Kontrollkästchen35",
    },
    "Sprint-Ziele erreicht": {0: "Kontrollkästchen37", 1: "Kontrollkästchen36"},
    "Arbeitsumfang angemessen": {  # vertikal -> invers (Skala 0/5/10/15)
        0: "Kontrollkästchen41", 5: "Kontrollkästchen40",
        10: "Kontrollkästchen39", 15: "Kontrollkästchen38",
    },
    # --- Prozessqualität (Seite 2, vertikal -> invers) ---
    "GitLab-Nutzung (Issues, Board)": {
        0: "Kontrollkästchen44", 1: "Kontrollkästchen43", 2: "Kontrollkästchen42",
    },
    "Branching-Workflow (Feature-Branches, MR)": {0: "Kontrollkästchen46", 1: "Kontrollkästchen45"},
    "Code-Reviews durchgefuehrt": {0: "Kontrollkästchen48", 1: "Kontrollkästchen47"},
    "Team-Organisation & Kommunikation": {
        0: "Kontrollkästchen51", 1: "Kontrollkästchen50", 2: "Kontrollkästchen49",
    },
    "Selbstständigkeit (proaktiv mit Tutor)": {
        0: "Kontrollkästchen54", 1: "Kontrollkästchen53", 2: "Kontrollkästchen52",
    },
}

# Header-Textfelder (aus der Template-Geometrie verifiziert)
FIELD_PRUEFER = "Textfeld0"   # "Prüfer"
FIELD_TEAM = "Textfeld1"      # "Geprüftes Team"
FIELD_TOTAL = "Textfeld2"     # "Gesamtpunktzahl" (Gesamtsumme)
PRUEFER_NAME = "Maximilian Fiegen"


def _section_crits(section_name):
    """Kriterien-Liste einer CATEGORIES-Sektion (oder [] + Warnung bei Drift)."""
    for name, crits in ev.CATEGORIES:
        if name == section_name:
            return list(crits)
    print(f"WARN: CATEGORIES-Sektion '{section_name}' nicht gefunden (fill_pdf-Drift?)",
          file=sys.stderr)
    return []


# Zwischensummen-Felder je Abschnitt -> Kriterien. Aus der Template-Geometrie:
# jedes "Punkte:"-Feld sitzt auf der Kopfzeile seines Abschnitts. Der Abschnitt
# "Code-Qualität" hat im Template KEIN Summen-Feld und wird daher uebersprungen.
# Die PDF-"Prozessqualität" buendelt die automatischen UND die manuellen Kriterien.
_MANUAL_CRITS = [name for _, crits in ev.MANUAL_CRITERIA for name, _ in crits]
SECTION_SUBTOTAL_FIELDS = {
    "Textfeld3": _section_crits("Sprintdokumentation"),
    "Textfeld13": _section_crits("Implementierte Funktionalität"),
    "Textfeld17": _section_crits("Prozessqualität (teilweise automatisierbar)") + _MANUAL_CRITS,
}


def read_xlsx_scores(xlsx_path):
    """Liest die Scores aus einer Bewertungs-xlsx (Spalte F = 'Deine Bewertung').

    Returns ({criterion: {"score": x, "note": y}}, total). 'note' wird mitgelesen,
    aber bewusst NICHT ins PDF geschrieben (Rueckwaerts-Kompatibilitaet).
    total = Summe aller numerischen Scores (= Gesamtpunktzahl).
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Bewertung"] if "Bewertung" in wb.sheetnames else wb.active
    out = {}
    total = 0
    for row in range(6, 60):
        crit = ws.cell(row=row, column=ev.COL_CRITERION).value
        if not crit or "umme" in str(crit) or "GESAMT" in str(crit):
            continue
        score = ws.cell(row=row, column=ev.COL_SCORE).value
        note = ws.cell(row=row, column=ev.COL_NOTE).value
        if isinstance(score, (int, float)):
            total += int(score)
        out[str(crit).strip()] = {"score": score, "note": note}
    return out, total


def _subtotal(scores, crits):
    """Summe der numerischen Scores ueber die gegebenen Kriterien."""
    s = 0
    for c in crits:
        v = scores.get(c, {}).get("score")
        if isinstance(v, (int, float)):
            s += int(v)
    return s


def _checkbox_on_states(reader):
    """{checkbox_field_name: on_state} aus dem PDF (z. B. '/Yes'), robust statt
    hartem '/Yes'. on_state = der nicht-'/Off'-Zustand des Kaestchens."""
    states = {}
    for name, f in (reader.get_fields() or {}).items():
        if f.get("/FT") != "/Btn":
            continue
        st = f.get("/_States_") or []
        on = next((s for s in st if s != "/Off"), "/Yes")
        states[str(name)] = on
    return states


def fill_pdf(template_path, scores, total_score, output_path, team_name=""):
    """Befuellt das PDF-Template (Checkboxen + Team + Pruefer + Summen)."""
    from pypdf import PdfReader, PdfWriter

    reader = PdfReader(str(template_path))
    writer = PdfWriter(clone_from=reader)
    # NeedAppearances, damit Viewer die Textwerte rendern. clone_from hat die
    # AcroForm bereits korrekt mitkopiert - die Reader-AcroForm NICHT erneut
    # hineinkopieren (erzeugt cross-document refs -> kaputter Feldbaum).
    try:
        writer.set_need_appearances_writer(True)
    except Exception:
        pass

    on_states = _checkbox_on_states(reader)

    updates = {}
    # Kontrollkaestchen zur jeweiligen Punktzahl
    for crit, data in scores.items():
        score = data.get("score")
        if not isinstance(score, (int, float)):
            continue  # "x"/leer = noch nicht bewertet -> kein Kreuz
        cb_map = CHECKBOX_MAP.get(crit)
        if cb_map and int(score) in cb_map:
            field = cb_map[int(score)]
            updates[field] = on_states.get(field, "/Yes")

    # Header: Team, Pruefer, Gesamtpunktzahl
    if team_name:
        updates[FIELD_TEAM] = str(team_name)
    updates[FIELD_PRUEFER] = PRUEFER_NAME
    updates[FIELD_TOTAL] = str(total_score)

    # Zwischensummen je Abschnitt
    for field, crits in SECTION_SUBTOTAL_FIELDS.items():
        if crits:
            updates[field] = str(_subtotal(scores, crits))

    for page in writer.pages:
        writer.update_page_form_field_values(page, updates, auto_regenerate=False)

    with open(output_path, "wb") as f:
        writer.write(f)


def main_for(local_folder):
    """Wiederverwendbar fuer Aufruf aus run_all.py."""
    mapping = json.loads((ev.OUTPUTS / "team_mapping.json").read_text(encoding="utf-8"))
    entry = next((m for m in mapping if m["local_folder"] == local_folder), None)
    if not entry:
        raise RuntimeError(f"Kein Mapping fuer {local_folder}")

    xlsx = ev.TEAMS / local_folder / f"Bewertung_{local_folder}.xlsx"
    if not xlsx.exists():
        raise RuntimeError(f"Keine Excel fuer {local_folder}")

    template = ev.BASE / "assets" / "Templates" / "Template Artifacts Exam Checklist Fillable.pdf"
    if not template.exists():
        raise RuntimeError(f"PDF-Template nicht gefunden: {template}")
    output = ev.TEAMS / local_folder / f"Bewertung_{local_folder}.pdf"

    team_name = entry.get("name") or local_folder
    scores, total = read_xlsx_scores(xlsx)
    fill_pdf(template, scores, total, output, team_name=team_name)
    print(f"   PDF: {output.name}")
    return output


def main():
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <local_team_folder>")
        sys.exit(2)
    out = main_for(sys.argv[1])
    print(f"OK: {out}")


if __name__ == "__main__":
    main()
